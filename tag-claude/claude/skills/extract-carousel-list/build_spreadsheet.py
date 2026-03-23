# /// script
# requires-python = ">=3.13"
# dependencies = [
#   "openpyxl",
# ]
# ///
# ABOUTME: Builds a styled Excel spreadsheet from a carousel list analysis
# ABOUTME: JSON file, with numbered rows, alternating colors, and a notes sheet.
"""
Build a styled Excel spreadsheet from a carousel list analysis JSON file.

Input JSON format:
{
  "list_type": "books",
  "columns": ["Title", "Author", "Description"],
  "items": [{"Title": "...", "Author": "...", "Description": "..."}, ...],
  "notes": "optional notes about the extraction",
  "source_directory": "2026-03-22",
  "image_count": 9
}
"""

from __future__ import annotations

import argparse
import json
import sys
from pathlib import Path

import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

# ── style constants ──────────────────────────────────────────────
HEADER_FILL = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
HEADER_FONT = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
DATA_FONT = Font(name="Calibri", size=11)
WRAP = Alignment(wrap_text=True, vertical="top")
THIN_BORDER = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)
ALT_ROW_FILL = PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid")


def estimate_column_width(header: str, values: list[str]) -> float:
    """Estimate column width from header and data values."""
    max_len = len(header)
    for v in values:
        # For multi-line values, use the longest single line
        for line in str(v).split("\n"):
            max_len = max(max_len, len(line))
    return min(60.0, max(15.0, max_len + 4))


def build_workbook(analysis: dict[str, object]) -> openpyxl.Workbook:
    """Build a styled workbook from the analysis dict."""
    list_type: str = str(analysis.get("list_type", "Items"))
    columns: list[str] = analysis["columns"]  # type: ignore[assignment]
    items: list[dict[str, str]] = analysis["items"]  # type: ignore[assignment]
    notes: str = str(analysis.get("notes", ""))
    source_directory: str = str(analysis.get("source_directory", ""))
    image_count: int = int(analysis.get("image_count", 0))

    sheet_title = list_type.replace("_", " ").title()

    wb = openpyxl.Workbook()
    ws = wb.active
    assert ws is not None
    ws.title = sheet_title

    # All columns: # + user-defined columns
    all_headers = ["#"] + columns
    all_col_widths: list[float] = [6.0]

    for col_name in columns:
        col_values = [str(item.get(col_name, "")) for item in items]
        all_col_widths.append(estimate_column_width(col_name, col_values))

    # Write headers
    for col_idx, (header, width) in enumerate(zip(all_headers, all_col_widths), start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = WRAP
        cell.border = THIN_BORDER
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    # Write data rows
    for row_idx, item in enumerate(items, start=2):
        is_alt = (row_idx - 2) % 2 == 1

        # Row number
        num_cell = ws.cell(row=row_idx, column=1, value=row_idx - 1)
        num_cell.font = DATA_FONT
        num_cell.alignment = Alignment(horizontal="center", vertical="top")
        num_cell.border = THIN_BORDER
        if is_alt:
            num_cell.fill = ALT_ROW_FILL

        # Data columns
        for col_idx, col_name in enumerate(columns, start=2):
            value = item.get(col_name, "")
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.font = DATA_FONT
            cell.alignment = WRAP
            cell.border = THIN_BORDER
            if is_alt:
                cell.fill = ALT_ROW_FILL

    # Freeze header row
    ws.freeze_panes = "A2"

    # Auto-filter
    if items:
        last_col = get_column_letter(len(all_headers))
        ws.auto_filter.ref = f"A1:{last_col}{len(items) + 1}"

    # ── Notes sheet ──────────────────────────────────────────────
    ws2 = wb.create_sheet("Notes")
    ws2.column_dimensions["A"].width = 18
    ws2.column_dimensions["B"].width = 80

    note_rows = [
        ("List Type", sheet_title),
        ("Items Found", str(len(items))),
        ("Source Directory", source_directory),
        ("Images Processed", str(image_count)),
        ("", ""),
        ("Notes", notes),
    ]
    for i, (label, value) in enumerate(note_rows, start=1):
        label_cell = ws2.cell(row=i, column=1, value=label)
        label_cell.font = Font(name="Calibri", bold=True, size=11)
        value_cell = ws2.cell(row=i, column=2, value=value)
        value_cell.font = Font(name="Calibri", size=11)
        value_cell.alignment = WRAP

    return wb


def main() -> None:
    parser = argparse.ArgumentParser(description="Build Excel spreadsheet from analysis JSON")
    parser.add_argument("input", help="Path to analysis JSON file")
    parser.add_argument("-o", "--output", help="Output .xlsx path (default: based on input name)")
    args = parser.parse_args()

    input_path = Path(args.input).resolve()
    if not input_path.exists():
        print(f"Error: File not found: {input_path}", file=sys.stderr)
        sys.exit(1)

    analysis = json.loads(input_path.read_text())

    for key in ("columns", "items"):
        if key not in analysis:
            print(f"Error: Missing required key '{key}' in JSON", file=sys.stderr)
            sys.exit(1)

    if not isinstance(analysis["columns"], list) or not analysis["columns"]:
        print("Error: 'columns' must be a non-empty list", file=sys.stderr)
        sys.exit(1)

    if args.output:
        output_path = Path(args.output).resolve()
    else:
        stem = input_path.stem.removesuffix("_analysis")
        output_path = input_path.parent.joinpath(stem + ".xlsx")

    wb = build_workbook(analysis)
    wb.save(str(output_path))
    print(f"Wrote {output_path} ({len(analysis['items'])} rows)")


if __name__ == "__main__":
    main()
