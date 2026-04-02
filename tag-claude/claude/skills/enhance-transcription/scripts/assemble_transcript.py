#!/usr/bin/env python3
"""
Assemble corrected transcript chunks and uncertainty data into final outputs.

After the LLM processes each chunk in Pass 2, it produces corrected chunk files.
This script:
  1. Reads all corrected chunk files from a directory
  2. Deduplicates overlapping utterances (keeps the later chunk's version,
     which benefits from more accumulated context)
  3. Writes the final corrected transcript
  4. Collects all uncertainty entries and writes them to a formatted Excel file

Usage:
    python assemble_transcript.py \
        --corrected-dir ./pass2_corrected/ \
        --output-transcript corrected_transcript.txt \
        --output-uncertainties uncertainties.xlsx \
        [--manifest ./chunks/manifest.json]

Expected format in each corrected chunk file:

    Corrected utterance lines:
        [HH:MM:SS] {utterance_NNN} Speaker Name: text...

    Uncertainty lines (only for LOW and UNIDENTIFIED):
        UNCERTAINTY|HH:MM:SS|utterance_NNN|Speaker X|Assigned Name|LOW|reason|text preview

The {utterance_NNN} tag is the global index from the preprocessor, used for
deduplication of overlapping chunks. If absent, deduplication falls back to
timestamp + text matching.
"""

from __future__ import annotations

import argparse
import json
import re
import sys
from dataclasses import dataclass
from pathlib import Path

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


# ---------------------------------------------------------------------------
# Data structures
# ---------------------------------------------------------------------------

@dataclass
class CorrectedUtterance:
    """A single corrected utterance ready for the final transcript."""
    timestamp: str
    utterance_index: int | None  # global index from preprocessor, if available
    speaker_name: str
    text: str
    source_chunk: int

    def format_output(self) -> str:
        """Format for the final transcript file (no index tag)."""
        return f"[{self.timestamp}] {self.speaker_name}: {self.text}"

    @property
    def sort_key(self) -> tuple[int, ...]:
        """Sortable representation of the timestamp."""
        parts = self.timestamp.split(":")
        return tuple(int(p) for p in parts)


@dataclass
class UncertaintyEntry:
    """An uncertain speaker identification for the review log."""
    timestamp: str
    utterance_index: int | None
    original_label: str
    assigned_name: str
    confidence: str  # LOW or UNIDENTIFIED
    reason: str
    text_preview: str
    source_chunk: int


# ---------------------------------------------------------------------------
# Parsing patterns
# ---------------------------------------------------------------------------

# Corrected utterance: [HH:MM:SS] {utterance_42} Speaker Name: text...
# The {utterance_NNN} tag is optional
CORRECTED_PATTERN = re.compile(
    r"^\[(\d{1,2}:\d{2}:\d{2})\]"
    r"(?:\s+\{utterance_(\d+)\})?"
    r"\s+(.+?):\s+(.*)$"
)

# Uncertainty: UNCERTAINTY|timestamp|utterance_NNN|orig_label|assigned|confidence|reason|preview
UNCERTAINTY_PATTERN = re.compile(
    r"^UNCERTAINTY\|([^|]*)\|([^|]*)\|([^|]*)\|([^|]*)\|([^|]*)\|([^|]*)\|(.*)$"
)


# ---------------------------------------------------------------------------
# Parsing
# ---------------------------------------------------------------------------

def parse_corrected_chunk(
    filepath: Path,
    chunk_index: int,
) -> tuple[list[CorrectedUtterance], list[UncertaintyEntry]]:
    """Parse a single corrected chunk file produced by the LLM in Pass 2."""
    utterances: list[CorrectedUtterance] = []
    uncertainties: list[UncertaintyEntry] = []

    for line in filepath.read_text(encoding="utf-8").splitlines():
        line = line.strip()
        if not line or line.startswith("#"):
            continue

        # Try uncertainty line first (starts with UNCERTAINTY|)
        if line.startswith("UNCERTAINTY|"):
            umatch = UNCERTAINTY_PATTERN.match(line)
            if umatch:
                utt_idx_str = umatch.group(2)
                utt_idx = int(utt_idx_str) if utt_idx_str.isdigit() else None
                uncertainties.append(UncertaintyEntry(
                    timestamp=umatch.group(1),
                    utterance_index=utt_idx,
                    original_label=umatch.group(3),
                    assigned_name=umatch.group(4),
                    confidence=umatch.group(5),
                    reason=umatch.group(6),
                    text_preview=umatch.group(7),
                    source_chunk=chunk_index,
                ))
            continue

        # Try corrected utterance line
        cmatch = CORRECTED_PATTERN.match(line)
        if cmatch:
            utt_idx_str = cmatch.group(2)
            utt_idx = int(utt_idx_str) if utt_idx_str else None
            utterances.append(CorrectedUtterance(
                timestamp=cmatch.group(1),
                utterance_index=utt_idx,
                speaker_name=cmatch.group(3),
                text=cmatch.group(4),
                source_chunk=chunk_index,
            ))

    return utterances, uncertainties


def load_all_chunks(
    corrected_dir: Path,
) -> tuple[list[CorrectedUtterance], list[UncertaintyEntry]]:
    """Load all corrected chunk files from a directory."""
    all_utterances: list[CorrectedUtterance] = []
    all_uncertainties: list[UncertaintyEntry] = []

    chunk_files = sorted(corrected_dir.glob("chunk_*.txt"))
    if not chunk_files:
        print(f"WARNING: No chunk_*.txt files found in {corrected_dir}", file=sys.stderr)
        return all_utterances, all_uncertainties

    print(f"Found {len(chunk_files)} corrected chunk files")

    for filepath in chunk_files:
        # Extract chunk index from filename (chunk_001.txt -> 1)
        stem = filepath.stem
        try:
            chunk_index = int(stem.split("_")[1])
        except (IndexError, ValueError):
            chunk_index = 0

        utterances, uncertainties = parse_corrected_chunk(filepath, chunk_index)
        all_utterances.extend(utterances)
        all_uncertainties.extend(uncertainties)
        print(f"  {filepath.name}: {len(utterances)} utterances, "
              f"{len(uncertainties)} uncertainties")

    return all_utterances, all_uncertainties


# ---------------------------------------------------------------------------
# Deduplication
# ---------------------------------------------------------------------------

def deduplicate_utterances(
    utterances: list[CorrectedUtterance],
) -> list[CorrectedUtterance]:
    """Remove duplicate utterances from overlapping chunks.

    When chunks overlap, the same utterance appears twice. We keep the version
    from the LATER chunk, since it had more accumulated speaker-map context.

    Dedup key: utterance_index if available, else timestamp + first 50 chars of text.
    """
    seen: dict[str | int, CorrectedUtterance] = {}

    for utt in utterances:
        if utt.utterance_index is not None:
            key: str | int = utt.utterance_index
        else:
            key = f"{utt.timestamp}|{utt.text[:50]}"

        existing = seen.get(key)
        if existing is None or utt.source_chunk > existing.source_chunk:
            seen[key] = utt

    # Sort by utterance_index if available, else by timestamp
    deduped = sorted(
        seen.values(),
        key=lambda u: (u.utterance_index if u.utterance_index is not None else 999999,
                       u.sort_key),
    )
    return deduped


def deduplicate_uncertainties(
    uncertainties: list[UncertaintyEntry],
) -> list[UncertaintyEntry]:
    """Remove duplicate uncertainty entries from overlapping chunks."""
    seen: dict[str | int, UncertaintyEntry] = {}

    for unc in uncertainties:
        if unc.utterance_index is not None:
            key: str | int = unc.utterance_index
        else:
            key = f"{unc.timestamp}|{unc.original_label}"

        existing = seen.get(key)
        if existing is None or unc.source_chunk > existing.source_chunk:
            seen[key] = unc

    return sorted(
        seen.values(),
        key=lambda u: (u.utterance_index if u.utterance_index is not None else 999999,
                       tuple(int(p) for p in u.timestamp.split(":"))),
    )


# ---------------------------------------------------------------------------
# Output: corrected transcript
# ---------------------------------------------------------------------------

def write_corrected_transcript(
    utterances: list[CorrectedUtterance],
    output_path: Path,
) -> None:
    """Write the final corrected transcript as a plain text file."""
    lines = [utt.format_output() for utt in utterances]
    output_path.write_text("\n\n".join(lines) + "\n", encoding="utf-8")
    print(f"\nCorrected transcript: {output_path}")
    print(f"  {len(utterances)} utterances")


# ---------------------------------------------------------------------------
# Output: uncertainty log Excel
# ---------------------------------------------------------------------------

FILL_LOW = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
FILL_UNIDENTIFIED = PatternFill(
    start_color="FF9999", end_color="FF9999", fill_type="solid"
)
HEADER_FILL = PatternFill(
    start_color="4472C4", end_color="4472C4", fill_type="solid"
)
HEADER_FONT = Font(name="Arial", bold=True, color="FFFFFF", size=11)
BODY_FONT = Font(name="Arial", size=10)


def write_uncertainty_xlsx(
    uncertainties: list[UncertaintyEntry],
    output_path: Path,
) -> None:
    """Write the uncertainty log as a formatted Excel file."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Uncertainty Log"

    headers = [
        "Timestamp",
        "Original Label",
        "Assigned Name",
        "Confidence",
        "Reason",
        "Text Preview",
    ]

    # Header row
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center")

    # Data rows
    for row_idx, unc in enumerate(uncertainties, start=2):
        values = [
            unc.timestamp,
            unc.original_label,
            unc.assigned_name,
            unc.confidence,
            unc.reason,
            unc.text_preview,
        ]
        fill = FILL_LOW if unc.confidence == "LOW" else FILL_UNIDENTIFIED

        for col, value in enumerate(values, start=1):
            cell = ws.cell(row=row_idx, column=col, value=value)
            cell.font = BODY_FONT
            cell.fill = fill

    # Column widths
    col_widths = [12, 16, 28, 16, 45, 65]
    for col, width in enumerate(col_widths, start=1):
        ws.column_dimensions[get_column_letter(col)].width = width

    # Freeze header row
    ws.freeze_panes = "A2"

    # Auto-filter
    if uncertainties:
        ws.auto_filter.ref = (
            f"A1:{get_column_letter(len(headers))}{len(uncertainties) + 1}"
        )

    wb.save(output_path)

    low_count = sum(1 for u in uncertainties if u.confidence == "LOW")
    unid_count = sum(1 for u in uncertainties if u.confidence == "UNIDENTIFIED")

    print(f"\nUncertainty log: {output_path}")
    print(f"  {len(uncertainties)} entries (LOW: {low_count}, UNIDENTIFIED: {unid_count})")


# ---------------------------------------------------------------------------
# Summary statistics
# ---------------------------------------------------------------------------

def print_summary(
    utterances: list[CorrectedUtterance],
    uncertainties: list[UncertaintyEntry],
) -> None:
    """Print a summary of the correction results."""
    # Unique speakers
    speakers = set()
    unidentified_count = 0
    for utt in utterances:
        if utt.speaker_name == "[UNIDENTIFIED]":
            unidentified_count += 1
        else:
            speakers.add(utt.speaker_name)

    print(f"\n{'='*60}")
    print(f"CORRECTION SUMMARY")
    print(f"{'='*60}")
    print(f"Total utterances:     {len(utterances)}")
    print(f"Unique speakers:      {len(speakers)}")
    print(f"Unidentified:         {unidentified_count} utterances")
    print(f"Uncertainties logged: {len(uncertainties)}")
    print(f"\nIdentified speakers:")
    for name in sorted(speakers):
        count = sum(1 for u in utterances if u.speaker_name == name)
        print(f"  {name}: {count} utterances")


# ---------------------------------------------------------------------------
# CLI
# ---------------------------------------------------------------------------

def main() -> None:
    parser = argparse.ArgumentParser(
        description="Assemble corrected transcript chunks into final outputs.",
    )
    parser.add_argument(
        "--corrected-dir", required=True, type=Path,
        help="Directory containing corrected chunk files (chunk_001.txt, ...)",
    )
    parser.add_argument(
        "--output-transcript", type=Path,
        default=Path("corrected_transcript.txt"),
        help="Output path for final corrected transcript (default: corrected_transcript.txt)",
    )
    parser.add_argument(
        "--output-uncertainties", type=Path,
        default=Path("uncertainties.xlsx"),
        help="Output path for uncertainty log Excel (default: uncertainties.xlsx)",
    )

    args = parser.parse_args()

    if not args.corrected_dir.exists():
        print(f"ERROR: Corrected directory not found: {args.corrected_dir}",
              file=sys.stderr)
        sys.exit(1)

    # Load
    all_utterances, all_uncertainties = load_all_chunks(args.corrected_dir)

    if not all_utterances:
        print("ERROR: No utterances found in corrected chunks.", file=sys.stderr)
        sys.exit(1)

    # Deduplicate
    print(f"\nDeduplicating {len(all_utterances)} utterances...")
    deduped_utterances = deduplicate_utterances(all_utterances)
    removed = len(all_utterances) - len(deduped_utterances)
    print(f"  {len(deduped_utterances)} unique ({removed} overlap duplicates removed)")

    print(f"Deduplicating {len(all_uncertainties)} uncertainty entries...")
    deduped_uncertainties = deduplicate_uncertainties(all_uncertainties)
    removed = len(all_uncertainties) - len(deduped_uncertainties)
    print(f"  {len(deduped_uncertainties)} unique ({removed} duplicates removed)")

    # Write outputs
    write_corrected_transcript(deduped_utterances, args.output_transcript)
    write_uncertainty_xlsx(deduped_uncertainties, args.output_uncertainties)

    # Summary
    print_summary(deduped_utterances, deduped_uncertainties)


if __name__ == "__main__":
    main()
