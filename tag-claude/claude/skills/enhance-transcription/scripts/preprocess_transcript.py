#!/usr/bin/env python3
"""
Transcript Preprocessor for the Enhance Transcription Skill.

This script prepares legislative hearing transcripts for LLM-based speaker
identification by:

1. Parsing the speaker-labeled transcript text and (optionally) the AssemblyAI JSON
2. Extracting low-confidence segments from the JSON for the LLM to flag
3. Parsing the speaker order list and committee members list from Excel
4. Splitting the transcript into overlapping chunks suitable for LLM context windows
5. Writing chunk files and a manifest that the LLM processes sequentially

Usage:
    python preprocess_transcript.py \
        --transcript speakers.txt \
        --speakers speakers.xlsx \
        --members members.xlsx \
        --output-dir ./chunks \
        [--json transcript.json] \
        [--chunk-size 200] \
        [--overlap 20]

Inputs:
    --transcript    Speaker-labeled transcript (.txt), lines like:
                    [HH:MM:SS] Speaker A: text...
    --json          (Optional) AssemblyAI JSON with word-level confidence scores
    --speakers      Speaker order list (.xlsx) with columns:
                    Session, #, Name, Title, Organization
    --members       Committee members list (.xlsx) with columns:
                    Name, Note, District, Party
    --output-dir    Directory to write chunk files into (created if needed)
    --chunk-size    Target utterances per chunk (default: 200)
    --overlap       Overlap utterances between chunks (default: 20)

Outputs in --output-dir:
    manifest.json           Chunking metadata and speaker lists for the LLM
    chunk_001.txt           First chunk of the transcript
    chunk_002.txt           Second chunk (with overlap from first), etc.
    low_confidence.json     (If --json provided) Segments with low confidence scores
"""

from __future__ import annotations

import argparse
import json
import re
import sys
from dataclasses import dataclass, field, asdict
from pathlib import Path

import openpyxl


# ---------------------------------------------------------------------------
# Data classes
# ---------------------------------------------------------------------------

@dataclass
class Utterance:
    """A single speaker-labeled utterance from the transcript."""
    line_number: int
    timestamp: str
    speaker_label: str
    text: str
    # Optional fields populated from JSON
    start_ms: int | None = None
    end_ms: int | None = None
    confidence: float | None = None


@dataclass
class Speaker:
    """A speaker from the sign-up sheet."""
    session: str
    order: int | None
    name: str
    title: str | None
    organization: str | None


@dataclass
class CommitteeMember:
    """A committee member."""
    name: str
    note: str | None  # e.g. "Chair", "Vice Chair"
    district: int | None
    party: str | None


@dataclass
class LowConfidenceSegment:
    """A segment with low diarization or transcription confidence."""
    utterance_index: int
    timestamp: str
    speaker_label: str
    confidence: float
    text_preview: str
    reason: str  # "low_utterance_confidence" or "low_word_confidence"


@dataclass
class ChunkInfo:
    """Metadata about a single chunk."""
    chunk_number: int
    filename: str
    start_utterance: int  # index into full utterance list
    end_utterance: int    # exclusive
    overlap_from_previous: int  # how many utterances overlap with prior chunk
    utterance_count: int
    time_range: str  # e.g. "00:00:00 - 01:23:45"


@dataclass
class Manifest:
    """Top-level manifest written for the LLM to consume."""
    total_utterances: int
    total_chunks: int
    chunk_size: int
    overlap: int
    speakers: list[dict]
    committee_members: list[dict]
    chunks: list[dict]
    has_low_confidence_data: bool
    unique_speaker_labels: list[str]
    speaker_label_frequencies: dict[str, int]


# ---------------------------------------------------------------------------
# Parsing functions
# ---------------------------------------------------------------------------

# Matches lines like: [00:00:00] Speaker A: text...
UTTERANCE_PATTERN = re.compile(
    r"^\[(\d{1,2}:\d{2}:\d{2})\]\s+(Speaker\s+\w+):\s+(.+)$"
)


def parse_transcript_text(filepath: Path) -> list[Utterance]:
    """Parse a speaker-labeled transcript text file into Utterance objects."""
    utterances: list[Utterance] = []
    lines = filepath.read_text(encoding="utf-8").splitlines()

    for line_num, line in enumerate(lines, start=1):
        line = line.strip()
        if not line:
            continue

        match = UTTERANCE_PATTERN.match(line)
        if match:
            timestamp, speaker_label, text = match.groups()
            utterances.append(Utterance(
                line_number=line_num,
                timestamp=timestamp,
                speaker_label=speaker_label,
                text=text,
            ))

    return utterances


def enrich_from_json(
    utterances: list[Utterance],
    json_path: Path,
    low_confidence_threshold: float = 0.70,
    low_word_confidence_threshold: float = 0.40,
    low_word_fraction_threshold: float = 0.20,
) -> list[LowConfidenceSegment]:
    """
    Enrich utterances with confidence data from AssemblyAI JSON and extract
    low-confidence segments.

    An utterance is flagged as low confidence if:
    - Its overall utterance confidence is below low_confidence_threshold, OR
    - More than low_word_fraction_threshold of its words have confidence below
      low_word_confidence_threshold

    Returns a list of LowConfidenceSegment for the LLM to review.
    """
    with open(json_path, encoding="utf-8") as f:
        data = json.load(f)

    json_utterances = data.get("utterances", [])
    low_confidence_segments: list[LowConfidenceSegment] = []

    # Build a lookup from (speaker_label, approximate_timestamp) to JSON utterance.
    # The text files use Speaker A/B/etc and the JSON uses just A/B/etc.
    # We match by index since both should be in the same order.
    for i, utt in enumerate(utterances):
        if i >= len(json_utterances):
            break

        json_utt = json_utterances[i]
        utt.start_ms = json_utt.get("start")
        utt.end_ms = json_utt.get("end")
        utt.confidence = json_utt.get("confidence")

        # Check utterance-level confidence
        if utt.confidence is not None and utt.confidence < low_confidence_threshold:
            low_confidence_segments.append(LowConfidenceSegment(
                utterance_index=i,
                timestamp=utt.timestamp,
                speaker_label=utt.speaker_label,
                confidence=utt.confidence,
                text_preview=utt.text[:120],
                reason="low_utterance_confidence",
            ))
            continue

        # Check word-level confidence
        words = json_utt.get("words", [])
        if words:
            low_conf_words = [
                w for w in words
                if w.get("confidence", 1.0) < low_word_confidence_threshold
            ]
            fraction_low = len(low_conf_words) / len(words)
            if fraction_low >= low_word_fraction_threshold:
                low_confidence_segments.append(LowConfidenceSegment(
                    utterance_index=i,
                    timestamp=utt.timestamp,
                    speaker_label=utt.speaker_label,
                    confidence=utt.confidence or 0.0,
                    text_preview=utt.text[:120],
                    reason=f"low_word_confidence ({len(low_conf_words)}/{len(words)} words below {low_word_confidence_threshold})",
                ))

    return low_confidence_segments


def parse_speakers_xlsx(filepath: Path) -> list[Speaker]:
    """Parse the speaker order list from an Excel file."""
    wb = openpyxl.load_workbook(filepath, read_only=True)
    ws = wb.active
    speakers: list[Speaker] = []

    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return speakers

    # Skip header row — expect: Session, #, Name, Title, Organization
    for row in rows[1:]:
        if len(row) < 3:
            continue
        session = str(row[0] or "").strip()
        order_raw = row[1]
        name = str(row[2] or "").strip()

        if not name:
            continue

        order: int | None = None
        if order_raw is not None:
            try:
                order = int(order_raw)
            except (ValueError, TypeError):
                order = None

        title = str(row[3]).strip() if len(row) > 3 and row[3] else None
        organization = str(row[4]).strip() if len(row) > 4 and row[4] else None

        speakers.append(Speaker(
            session=session,
            order=order,
            name=name,
            title=title,
            organization=organization,
        ))

    wb.close()
    return speakers


def parse_members_xlsx(filepath: Path) -> list[CommitteeMember]:
    """Parse the committee members list from an Excel file."""
    wb = openpyxl.load_workbook(filepath, read_only=True)
    ws = wb.active
    members: list[CommitteeMember] = []

    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return members

    # Skip header row — expect: Name, Note, District, Party
    for row in rows[1:]:
        if len(row) < 1:
            continue
        name = str(row[0] or "").strip()
        if not name:
            continue

        note = str(row[1]).strip() if len(row) > 1 and row[1] else None
        district: int | None = None
        if len(row) > 2 and row[2] is not None:
            try:
                district = int(row[2])
            except (ValueError, TypeError):
                district = None
        party = str(row[3]).strip() if len(row) > 3 and row[3] else None

        members.append(CommitteeMember(
            name=name,
            note=note,
            district=district,
            party=party,
        ))

    wb.close()
    return members


# ---------------------------------------------------------------------------
# Chunking
# ---------------------------------------------------------------------------

def find_chunk_boundary(
    utterances: list[Utterance],
    target_index: int,
    search_window: int = 15,
) -> int:
    """
    Find a natural chunk boundary near target_index by looking for testimony
    transitions: a "thank you" from one speaker followed by the chair calling
    the next speaker.

    Returns the best index to break at (the break happens BEFORE this index).
    """
    best = target_index
    best_score = 0

    start = max(0, target_index - search_window)
    end = min(len(utterances), target_index + search_window)

    for i in range(start, end):
        if i <= 0 or i >= len(utterances):
            continue

        prev_text = utterances[i - 1].text.lower()
        curr_text = utterances[i].text.lower()
        prev_speaker = utterances[i - 1].speaker_label
        curr_speaker = utterances[i].speaker_label

        score = 0

        # Speaker change at this boundary
        if prev_speaker != curr_speaker:
            score += 1

        # Previous utterance ends with thanks/closing
        if any(phrase in prev_text for phrase in [
            "thank you", "thanks", "appreciate", "that concludes",
        ]):
            score += 2

        # Current utterance starts with greeting/introduction
        if any(phrase in curr_text[:60] for phrase in [
            "good morning", "good afternoon", "good evening",
            "my name is", "thank you for the opportunity",
            "we have", "next we have", "is here",
        ]):
            score += 2

        if score > best_score:
            best_score = score
            best = i

    return best


def create_chunks(
    utterances: list[Utterance],
    chunk_size: int = 200,
    overlap: int = 20,
) -> list[ChunkInfo]:
    """
    Split utterances into overlapping chunks, breaking at natural boundaries.

    Returns a list of ChunkInfo describing each chunk's boundaries.
    """
    chunks: list[ChunkInfo] = []
    total = len(utterances)
    current_start = 0
    chunk_number = 1

    while current_start < total:
        # Target end for this chunk
        target_end = min(current_start + chunk_size, total)

        if target_end < total:
            # Find a natural boundary near the target end
            actual_end = find_chunk_boundary(utterances, target_end)
            # Don't let boundary search push us backwards too far
            if actual_end <= current_start + (chunk_size // 2):
                actual_end = target_end
        else:
            actual_end = total

        # Calculate overlap from previous chunk
        overlap_count = 0
        if chunk_number > 1 and chunks:
            prev_end = chunks[-1].end_utterance
            overlap_count = max(0, prev_end - current_start)

        time_start = utterances[current_start].timestamp
        time_end = utterances[actual_end - 1].timestamp

        chunks.append(ChunkInfo(
            chunk_number=chunk_number,
            filename=f"chunk_{chunk_number:03d}.txt",
            start_utterance=current_start,
            end_utterance=actual_end,
            overlap_from_previous=overlap_count,
            utterance_count=actual_end - current_start,
            time_range=f"{time_start} - {time_end}",
        ))

        # Next chunk starts overlap utterances before the end of this one
        next_start = actual_end - overlap
        if next_start <= current_start:
            # Avoid infinite loop — just move forward
            next_start = actual_end

        current_start = next_start
        chunk_number += 1

    return chunks


def write_chunk_file(
    utterances: list[Utterance],
    chunk: ChunkInfo,
    output_dir: Path,
) -> None:
    """Write a single chunk file with utterances formatted for the LLM."""
    chunk_utterances = utterances[chunk.start_utterance:chunk.end_utterance]

    lines: list[str] = []
    lines.append(f"# Chunk {chunk.chunk_number}")
    lines.append(f"# Time range: {chunk.time_range}")
    lines.append(f"# Utterances: {chunk.start_utterance} to {chunk.end_utterance - 1}")
    if chunk.overlap_from_previous > 0:
        lines.append(
            f"# NOTE: First {chunk.overlap_from_previous} utterances overlap "
            f"with previous chunk (for context continuity)"
        )
    lines.append("")

    for i, utt in enumerate(chunk_utterances):
        global_index = chunk.start_utterance + i
        confidence_note = ""
        if utt.confidence is not None and utt.confidence < 0.70:
            confidence_note = f"  [LOW CONFIDENCE: {utt.confidence:.2f}]"

        lines.append(
            f"[{utt.timestamp}] {{utterance_{global_index}}} "
            f"{utt.speaker_label}: {utt.text}{confidence_note}"
        )
        lines.append("")

    filepath = output_dir / chunk.filename
    filepath.write_text("\n".join(lines), encoding="utf-8")


# ---------------------------------------------------------------------------
# Speaker label analysis
# ---------------------------------------------------------------------------

def analyze_speaker_labels(utterances: list[Utterance]) -> dict[str, int]:
    """Count frequency of each speaker label across the full transcript."""
    frequencies: dict[str, int] = {}
    for utt in utterances:
        frequencies[utt.speaker_label] = frequencies.get(utt.speaker_label, 0) + 1
    return dict(sorted(frequencies.items(), key=lambda x: -x[1]))


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------

def main() -> None:
    parser = argparse.ArgumentParser(
        description="Preprocess a legislative hearing transcript for LLM-based speaker identification.",
    )
    parser.add_argument(
        "--transcript", required=True, type=Path,
        help="Speaker-labeled transcript (.txt)",
    )
    parser.add_argument(
        "--json", type=Path, default=None,
        help="(Optional) AssemblyAI JSON with word-level confidence",
    )
    parser.add_argument(
        "--speakers", required=True, type=Path,
        help="Speaker order list (.xlsx)",
    )
    parser.add_argument(
        "--members", required=True, type=Path,
        help="Committee members list (.xlsx)",
    )
    parser.add_argument(
        "--output-dir", required=True, type=Path,
        help="Directory to write chunk files and manifest",
    )
    parser.add_argument(
        "--chunk-size", type=int, default=200,
        help="Target utterances per chunk (default: 200)",
    )
    parser.add_argument(
        "--overlap", type=int, default=20,
        help="Overlap utterances between chunks (default: 20)",
    )

    args = parser.parse_args()

    # Validate inputs exist
    for path_arg, label in [
        (args.transcript, "Transcript"),
        (args.speakers, "Speaker list"),
        (args.members, "Members list"),
    ]:
        if not path_arg.exists():
            print(f"ERROR: {label} file not found: {path_arg}", file=sys.stderr)
            sys.exit(1)

    if args.json and not args.json.exists():
        print(f"ERROR: JSON file not found: {args.json}", file=sys.stderr)
        sys.exit(1)

    # Create output directory
    args.output_dir.mkdir(parents=True, exist_ok=True)

    # 1. Parse all inputs
    print(f"Parsing transcript: {args.transcript}")
    utterances = parse_transcript_text(args.transcript)
    print(f"  Found {len(utterances)} utterances")

    print(f"Parsing speaker list: {args.speakers}")
    speakers = parse_speakers_xlsx(args.speakers)
    print(f"  Found {len(speakers)} speakers")

    print(f"Parsing members list: {args.members}")
    members = parse_members_xlsx(args.members)
    print(f"  Found {len(members)} committee members")

    # 2. Enrich from JSON if provided
    low_confidence_segments: list[LowConfidenceSegment] = []
    if args.json:
        print(f"Enriching from JSON: {args.json}")
        low_confidence_segments = enrich_from_json(utterances, args.json)
        print(f"  Found {len(low_confidence_segments)} low-confidence segments")

    # 3. Analyze speaker labels
    label_frequencies = analyze_speaker_labels(utterances)
    unique_labels = list(label_frequencies.keys())
    print(f"Speaker labels found: {len(unique_labels)}")
    print(f"  Most frequent: {list(label_frequencies.items())[:5]}")

    # 4. Create chunks
    print(f"Creating chunks (size={args.chunk_size}, overlap={args.overlap})")
    chunks = create_chunks(utterances, args.chunk_size, args.overlap)
    print(f"  Created {len(chunks)} chunks")

    # 5. Write chunk files
    for chunk in chunks:
        write_chunk_file(utterances, chunk, args.output_dir)
        print(f"  Wrote {chunk.filename} ({chunk.utterance_count} utterances, {chunk.time_range})")

    # 6. Write low confidence data if available
    has_low_confidence = len(low_confidence_segments) > 0
    if has_low_confidence:
        lc_path = args.output_dir / "low_confidence.json"
        lc_data = [asdict(seg) for seg in low_confidence_segments]
        lc_path.write_text(json.dumps(lc_data, indent=2), encoding="utf-8")
        print(f"  Wrote low_confidence.json ({len(low_confidence_segments)} segments)")

    # 7. Write manifest
    manifest = Manifest(
        total_utterances=len(utterances),
        total_chunks=len(chunks),
        chunk_size=args.chunk_size,
        overlap=args.overlap,
        speakers=[asdict(s) for s in speakers],
        committee_members=[asdict(m) for m in members],
        chunks=[asdict(c) for c in chunks],
        has_low_confidence_data=has_low_confidence,
        unique_speaker_labels=unique_labels,
        speaker_label_frequencies=label_frequencies,
    )

    manifest_path = args.output_dir / "manifest.json"
    manifest_path.write_text(
        json.dumps(asdict(manifest), indent=2),
        encoding="utf-8",
    )
    print(f"  Wrote manifest.json")

    print(f"\nDone. Output directory: {args.output_dir}")
    print(f"Next step: Use the enhance-transcription skill to process chunks sequentially.")


if __name__ == "__main__":
    main()
